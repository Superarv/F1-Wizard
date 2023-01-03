from openpyxl import Workbook, load_workbook
import os

wb = load_workbook('f1Stats.xlsx', data_only=True)
ws = wb.active

columns = {
    "2023 Team" : "B",
    "Debut Race" : "C",
    "Debut Team" : "D",
    "WDCs" : "E",
    "Wins" : "F",
    "Podiums" : "G",
    "Poles" : "H",
    "Grand Slams" : "I",
    "Race Entries" : "J",
    "Highest WDC Standing" : "K",
    "Last Race" : "L",
    "Last Win" : "M",
    "DNFs" : "N",
    "Hatricks" : "O",
    "Current Teammate" : "P",
    "Birthplace" : "Q",
    "Nationality" : "R",
    "Birthday" : "S",
    "Fastest Laps" : "T",
    "Racing Number" : "U",
    "Most Wins In Season" : "V",
    "Seasons" : "W",
    "Career Points" : "X",
    "Win %" : "Y",
    "Grand Slam %" : "Z",
    "DNF %" : "AA",
    "WDC %" : "AB",
    "Birth Nation" : "AC"
}

rows = {
    "Max Verstappen" : "2",
    "Sergio Perez" : "3",
    "Lewis Hamilton" : "4",
    "George Russell" : "5",
    "Oscar Piastri" : "6",
    "Lando Norris" : "7",
    "Yuki Tsunoda" : "8",
    "Nyck De Vries" : "9",
    "Logan Sargeant" : "10",
    "Alexander Albon" : "11",
    "Pierre Gasly" : "12",
    "Esteban Ocon" : "13",
    "Lance Stroll" : "14",
    "Fernando Alonso" : "15",
    "Carlos Sainz" : "16",
    "Charles Leclerc" : "17",
    "Valtteri Bottas" : "18",
    "Zhou Guanyu" : "19",
    "Kevin Magnussen" : "20",
    "Nico Hulkenberg" : "21"
}

primaryColors = {
    "Red Bull" : "#F4801F",
    "Mercedes" : "#C6C6C6",
    "Ferrari" : "#A6051A",
    "Alfa Romeo" : "#981E32",
    "Alphatauri" : "#00293F",
    "Mclaren" : "#FF8000",
    "Williams Racing" : "#00A3E0",
    "Haas" : "#F62039",
    "Automobiles Alpine" : "#005BA9",
    "Aston Martin" : "#00352F",
}

secondaryColors = {
    "Red Bull" : "#000000",
    "Mercedes" : "#00A19C",
    "Ferrari" : "#FFEB00",
    "Alfa Romeo" : "#F9F9F8",
    "Alphatauri" : "#FFFFFF",
    "Mclaren" : "#47C7FC",
    "Williams Racing" : "#041E42",
    "Haas" : "#F9F2F2",
    "Automobiles Alpine" : "#FD4BC7",
    "Aston Martin" : "#CEDC00",
}

def findDriverStat(driver, stat):
    cell = columns[stat] + rows[driver]
    return ws[cell].value

def teamPrimaryColor(team):
    return primaryColors[team]
def teamSecondaryColor(team):
    return secondaryColors[team]

